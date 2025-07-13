import json
import threading
import logging
from datetime import date, datetime
from decimal import Decimal
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_http_methods


import requests
from django.shortcuts import render, redirect , get_object_or_404
from django.http import HttpResponseServerError
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import DetailView, UpdateView, ListView, CreateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.db.models import Q , Prefetch
from django.forms import inlineformset_factory
from django.utils.dateparse import parse_date
from django.urls import reverse
from django.utils import timezone
from django.contrib import messages
from django.conf import settings
from django.db import transaction
from django.core import serializers

from .models.dailyreport import DailyReport, Manpower, Equipment, Activity, Photo
from .models.data import Project, Client, Contact, Drawing
from .forms import ContactForm, ClientForm, DailyReportForm, ManpowerFormSet, EquipmentFormSet, ActivityFormSet, PhotoFormSet, DrawingFormSet

import io
import os
from collections import defaultdict

from openpyxl  import Workbook , load_workbook
from openpyxl.utils      import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles        import Font, Alignment, PatternFill

import re
import zipfile

from PIL import Image as PILImage, UnidentifiedImageError


from reportes.utils.send_mail import gas
from django.contrib.auth import logout

from django.http import JsonResponse
from django.views import View

logger = logging.getLogger(__name__)

# Weather API settings
GS_URL = 'https://script.google.com/macros/s/.../exec'
OWM_API_KEY = '6cddd2923580c11024597c9bb7bf5b55'

import os, io, re, zipfile
from collections import defaultdict
from datetime import datetime
from django.conf import settings
from django.db import transaction
from django.core import serializers
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from openpyxl import load_workbook

from .models import Project, DailyReport, Manpower, Equipment, Activity

import os
import re
import io
import zipfile
from collections import defaultdict
from datetime import datetime, date

from django.conf import settings
from django.core import serializers
from django.db import transaction
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from openpyxl import load_workbook

from reportes.models.data import Project
from reportes.models.dailyreport import DailyReport, Manpower, Equipment, Activity

# where we stash backups
BACKUP_ROOT = getattr(
    settings,
    "DAILYREPORT_BACKUP_ROOT",
    os.path.join(settings.BASE_DIR, "backups", "dailyreports")
)

# regex to detect YYYY-MM-DD sheet names
DATE_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})")

@require_http_methods(["GET", "POST"])
def import_daily_reports(request):
    """
    GET:
      - render upload form
      - pass `projects` (active) and `backups_json` for rollback UI

    POST:
      - if files uploaded → run import logic exactly as before
      - elif rollback requested → restore from chosen JSON backup
    """
    today = date.today()

    # 1) Handle POST
    if request.method == "POST":
        # a) Rollback branch
        if "backup" in request.POST and "project_id" in request.POST:
            pid    = request.POST["project_id"]
            backup = request.POST["backup"]
            backup_path = os.path.join(BACKUP_ROOT, pid, backup)
            if os.path.exists(backup_path):
                try:
                    with open(backup_path, "r", encoding="utf-8") as bf:
                        objs = list(serializers.deserialize("json", bf.read()))
                    count = 0
                    for des in objs:
                        des.save()
                        count += 1
                    messages.success(
                        request,
                        f"Restored {count} report(s) from backup “{backup}”."
                    )
                except Exception as e:
                    messages.error(
                        request,
                        f"Failed to restore from {backup}: {e}"
                    )
            else:
                messages.error(
                    request,
                    f"Backup file not found: {backup}"
                )
            return redirect("reportes:import_daily_reports")

        # b) Import branch
        successes = []
        errors    = []
        processed = defaultdict(set)

        def find_project_id(ws):
            for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
                if row[0] == "Project ID":
                    try:
                        return int(row[1])
                    except Exception:
                        return None
            return None

        def process_workbook(fp, filename):
            try:
                wb = load_workbook(filename=fp, data_only=True)
            except Exception as e:
                errors.append(f"{filename}: not a valid XLSX ({e})")
                return

            # 1) read Project ID from summary
            summary = wb.worksheets[0]
            proj_id = find_project_id(summary)
            if not proj_id:
                errors.append(f"{filename}: missing/invalid Project ID")
                return

            try:
                proj = Project.objects.get(pk=proj_id, end_date__gte=today)
            except Project.DoesNotExist:
                errors.append(f"{filename}: Project #{proj_id} not found or inactive")
                return

            # 2) for each detail sheet
            for sheet in wb.worksheets[1:]:
                title = sheet.title.strip()
                dr = None
                # date-named?
                m = DATE_RE.match(title)
                if m:
                    dt = datetime.strptime(m.group(1), "%Y-%m-%d").date()
                    dr = DailyReport.objects.filter(project=proj, date=dt).first()
                else:
                    m2 = re.match(r"Report\s*#(\d+)", title)
                    if m2:
                        dr = DailyReport.objects.filter(pk=int(m2.group(1))).first()

                if not dr:
                    errors.append(f"{filename} → sheet '{title}': skipped (no matching report)")
                    continue

                processed[proj.id].add(dr.id)

                # atomic wipe & reimport
                with transaction.atomic():
                    try:
                        dr.manpower.all().delete()
                        dr.equipment.all().delete()
                        dr.activities.all().delete()

                        rows = list(sheet.iter_rows(values_only=True))
                        def find_section(name):
                            for idx, row in enumerate(rows):
                                if row and row[0] and str(row[0]).strip() == name:
                                    return idx
                            return None

                        sections = [
                            ("Manpower",  Manpower,  ["role","quantity","hours","comments"]),
                            ("Equipment", Equipment, ["equipment","quantity","hours_used","comments"]),
                            ("Activities", Activity, ["activity","quantity","unit","description"]),
                        ]

                        for sec_name, model, fields in sections:
                            idx = find_section(sec_name)
                            if idx is None:
                                continue
                            r = idx + 2
                            while r < len(rows) and rows[r] and rows[r][0] is not None:
                                raw = rows[r][:len(fields)]
                                data = {}
                                for fname, val in zip(fields, raw):
                                    if val is None:
                                        # enforce non-null on comments
                                        val = "" if fname=="comments" else 0 if fname in ("quantity","hours","hours_used") else ""
                                    data[fname] = val
                                model.objects.create(daily_report=dr, **data)
                                r += 1

                        successes.append(
                            f"{filename} → project #{proj_id}, report {dr.date}: imported"
                        )
                    except Exception as e:
                        transaction.set_rollback(True)
                        errors.append(f"{filename} → sheet '{title}': ERROR {e}")

        # process each uploaded file
        for f in request.FILES.getlist("files"):
            name = f.name.lower()
            if name.endswith(".zip"):
                try:
                    zf = zipfile.ZipFile(f)
                except zipfile.BadZipFile:
                    errors.append(f"{f.name}: invalid ZIP")
                    continue
                for zi in zf.infolist():
                    if zi.filename.lower().endswith(".xlsx"):
                        data = zf.read(zi)
                        process_workbook(io.BytesIO(data), zi.filename)
                zf.close()
            elif name.endswith(".xlsx"):
                process_workbook(f, f.name)
            else:
                errors.append(f"{f.name}: unsupported, skipped")

        # backup & delete orphans
        for pid, seen in processed.items():
            orphan_qs = DailyReport.objects.filter(project_id=pid).exclude(pk__in=seen)
            if not orphan_qs.exists():
                continue
            backup_dir = os.path.join(BACKUP_ROOT, str(pid))
            os.makedirs(backup_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d%H%M%S")
            fname = f"proj_{pid}_orphans_{ts}.json"
            path  = os.path.join(backup_dir, fname)
            with open(path, "w", encoding="utf-8") as bf:
                bf.write(serializers.serialize("json", orphan_qs))
            # keep only last 10
            files = sorted(
                (os.path.join(backup_dir, fn) for fn in os.listdir(backup_dir) if fn.endswith(".json")),
                key=os.path.getmtime, reverse=True
            )
            for old in files[10:]:
                try: os.remove(old)
                except OSError: pass
            count = orphan_qs.count()
            orphan_qs.delete()
            messages.info(
                request,
                f"{count} old reports for project #{pid} backed up to {fname} and deleted"
            )

        # flash import results
        for m in successes: messages.success(request, m)
        for m in errors:    messages.error(request, m)

        return redirect("reportes:import_daily_reports")

    # 2) Handle GET
    # build list of active projects
    active_projects = Project.objects.filter(end_date__gte=today).order_by("name")
    # build a map of backups per project
    backups_map = {}
    for proj in active_projects:
        proj_dir = os.path.join(BACKUP_ROOT, str(proj.id))
        if not os.path.isdir(proj_dir):
            continue
        items = []
        for fn in sorted(os.listdir(proj_dir), reverse=True):
            if fn.endswith(".json"):
                full = os.path.join(proj_dir, fn)
                created = datetime.fromtimestamp(os.path.getmtime(full))\
                                  .strftime("%Y-%m-%d %H:%M")
                items.append({"filename": fn, "created": created})
        if items:
            backups_map[proj.id] = items

    return render(request, "reportes/import_reports.html", {
        "projects":     active_projects,
        "backups_json": json.dumps(backups_map),
    })

def build_wb(proj, request):
    """
    Return an openpyxl.Workbook for a single Project (with one 'Summary' sheet
    including project metadata + drawings, then one sheet per DailyReport).
    """
    wb     = Workbook()
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    gray   = PatternFill("solid", fgColor="DDDDDD")

    # — 1) Summary sheet —
    ws = wb.active
    safe_title = re.sub(r'[\\\/\?\*\[\]\:]', '_', proj.name)
    ws.title = safe_title[:31]
    
    row = 1

    # Project title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=f"{proj.name} ({proj.client})")
    c.font, c.alignment = bold, center
    row += 1

    # Project period
    period = f"Start: {proj.start_date or '-'}    End: {proj.end_date or '-'}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=period)
    c.font, c.alignment = bold, center
    row += 2

    # Project Details header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    hdr = ws.cell(row=row, column=1, value="Project Details")
    hdr.font, hdr.fill, hdr.alignment = bold, gray, center
    row += 1

    # Details rows (including Project ID)
    details = [
        ("Project ID",        proj.pk),
        ("Project Name",      proj.name),
        ("Client",            str(proj.client)),
        ("Location",          proj.location_city or "-"),
        ("Contract Number",   proj.contract_number or "-"),
        ("Notify Contacts",   ", ".join(
                                 f"{c.name} <{c.email}>"
                                 for c in proj.notify_contacts.all()
                               ) or "-"),
    ]
    for label, val in details:
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = bold
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 1

    # — Drawings —
    if proj.drawings.exists():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value="Project Drawings")
        c.font, c.fill, c.alignment = bold, gray, center
        row += 1

        col = 1
        for d in proj.drawings.all():
            cell_ref = f"{get_column_letter(col)}{row}"
            try:
                pil = PILImage.open(d.file.path)
                pil.thumbnail((120, 100), PILImage.LANCZOS)
                buf = io.BytesIO()
                pil.save(buf, format="PNG")
                buf.seek(0)
                img = XLImage(buf)
                img.anchor = cell_ref
                ws.add_image(img)
            except (UnidentifiedImageError, OSError):
                link_cell = ws.cell(
                    row=row, column=col,
                    value=f"Download {os.path.basename(d.file.name)}"
                )
                link_cell.hyperlink = request.build_absolute_uri(d.file.url)
                link_cell.font = Font(color="0000FF", underline="single")
            col += 2
            if col > 4:
                col = 1
                row += 7
        row += 8

    # — 2) One sheet per DailyReport —
    reports = (
        DailyReport.objects
        .filter(project=proj)
        .order_by("date")
        .prefetch_related("manpower", "equipment", "activities", "photos")
    )

    def write_section(ws, r, title, headers, items, attrs):
        # section header
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=len(headers))
        c = ws.cell(row=r, column=1, value=title)
        c.font, c.fill, c.alignment = bold, gray, center
        r += 1
        # column headers
        for i, h in enumerate(headers, start=1):
            hc = ws.cell(row=r, column=i, value=h)
            hc.font, hc.alignment = bold, center
        r += 1
        # data rows
        for obj in items:
            for i, attr in enumerate(attrs, start=1):
                ws.cell(row=r, column=i, value=getattr(obj, attr))
            r += 1
        return r + 1  # blank row after

    for dr in reports:
        ws = wb.create_sheet(title=dr.date.isoformat()[:31])
        r = 1

        # Report header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=f"Report #{dr.id} — {dr.date}")
        c.font, c.alignment = bold, center
        r += 2

        # Sections
        r = write_section(
            ws, r, "Manpower",
            ["Role","Qty","Hours","Comments"],
            dr.manpower.all(),
            ["role","quantity","hours","comments"]
        )
        r = write_section(
            ws, r, "Equipment",
            ["Equipment","Qty","Hours Used","Comments"],
            dr.equipment.all(),
            ["equipment","quantity","hours_used","comments"]
        )
        r = write_section(
            ws, r, "Activities",
            ["Activity","Qty","Unit","Description"],
            dr.activities.all(),
            ["activity","quantity","unit","description"]
        )

        # Photos
        if dr.photos.exists():
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value="Photos")
            c.font, c.fill, c.alignment = bold, gray, center
            r += 1
            for p in dr.photos.all():
                try:
                    pil = PILImage.open(p.photo.path)
                    pil.thumbnail((100, 100), PILImage.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, format="PNG")
                    buf.seek(0)
                    img = XLImage(buf)
                    img.anchor = f"A{r}"
                    ws.add_image(img)
                except (UnidentifiedImageError, OSError):
                    link_cell = ws.cell(
                        row=r, column=1,
                        value=f"Download {os.path.basename(p.photo.name)}"
                    )
                    link_cell.hyperlink = request.build_absolute_uri(p.photo.url)
                    link_cell.font = Font(color="0000FF", underline="single")
                r += 6

    # Remove default “Sheet” if left over
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Adjust column widths
    for sheet in wb.worksheets:
        for idx, w in enumerate([20, 10, 12, 40], start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = w

    return wb



@require_http_methods(["GET", "POST"])
def export_project_reports(request):
    today = timezone.localdate()

    if request.method == "POST":
        # 1) pull selected project IDs from the form
        ids = request.POST.getlist("projects")
        projects = Project.objects.filter(pk__in=ids, end_date__gte=today).order_by("name")

        if not projects:
            # no projects selected or all are inactive
            # you could add a message here if you use django.contrib.messages
            return redirect(request.path)

        # 2) single‐project → return one .xlsx
        if len(projects) == 1:
            proj = projects[0]
            wb = build_wb(proj, request)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            resp = HttpResponse(
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = f'attachment; filename="{proj.name}.xlsx"'
            return resp

        # 3) multiple → bundle into a ZIP
        zipbuf = io.BytesIO()
        with zipfile.ZipFile(zipbuf, "w") as zf:
            for proj in projects:
                wb = build_wb(proj, request)
                tmp = io.BytesIO()
                wb.save(tmp)
                zf.writestr(f"{proj.name}.xlsx", tmp.getvalue())
        zipbuf.seek(0)

        resp = HttpResponse(zipbuf.getvalue(), content_type="application/zip")
        resp["Content-Disposition"] = 'attachment; filename="Projects_Reports.zip"'
        return resp

    # GET: render the project-selection form (only active ones)
    active = Project.objects.filter(end_date__gte=today).order_by("name")
    return render(request, "reportes/export_reports.html", {
        "projects": active,
    })


@require_GET
def export_form(request):
    """GET → show only currently active projects for user to select."""
    today  = timezone.localdate()
    active = Project.objects.filter(end_date__gte=today).order_by("name")
    return render(request, "reportes/export_reports.html", {"projects": active})
class ProjectListJsonView(View):
    def get(self, request):
        client_id = request.GET.get('client')
        qs = Project.objects.filter(client_id=client_id, active=True) if client_id else Project.objects.none()
        data = [
            {
                "id": p.pk,
                "name": p.name,
                "contract_number": p.contract_number or ""
            }
            for p in qs.order_by('name')
        ]
        return JsonResponse(data, safe=False)

def logout_view(request):
    """
    Logs out the current user and redirects to the full report list.
    """
    logout(request)
    return redirect('reportes:full_dailyreport_list')

#@login_required
def daily_report_list(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    # pull in all reports for this project, newest first
    reports = project.dailyreport_set.order_by('-date').prefetch_related(
        'manpower', 'equipment', 'activities', 'photos'
    )
    return render(request, 'reportes/dailyreport_list.html', {
        'project': project,
        'reports': reports,
    })




#@method_decorator(login_required, name='dispatch')
class DailyReportDetailView(DetailView):
    model = DailyReport
    template_name = 'reportes/dailyreport_detail.html'
    context_object_name = 'report'



method_decorator(login_required, name='dispatch')
class DailyReportUpdateView(UpdateView):
    model = DailyReport
    form_class = DailyReportForm
    template_name = 'reportes/dailyreport_edit.html'
    success_url = reverse_lazy('reportes:dailyreport_list')  # or to detail/success

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        instance = self.object
        context['manpower_formset'] = ManpowerFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=instance
        )
        context['equipment_formset'] = EquipmentFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=instance
        )
        context['activity_formset'] = ActivityFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=instance
        )
        context['photo_formset'] = PhotoFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=instance
        )
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        # Save all formsets
        for fs_name in ('manpower_formset', 'equipment_formset', 'activity_formset', 'photo_formset'):
            formset = self.get_context_data()[fs_name]
            formset.instance = self.object
            formset.save()
        return response

@csrf_exempt
@login_required
def daily_report(request):
    if request.method == 'POST':
        try:
            # 1) parse incoming JSON payload
            payload = request.POST.get('payload')
            if not payload:
                raise ValueError("Missing payload JSON data")
            data = json.loads(payload)

            # 2) determine report date (fallback to today)
            raw_date = data.get('date')
            try:
                report_date = datetime.fromisoformat(raw_date).date() if raw_date else date.today()
            except Exception:
                logger.warning("Invalid date '%s', defaulting to today", raw_date)
                report_date = date.today()

            # 3) lookup project & its city
            project = Project.objects.get(pk=int(data['project']))
            city = project.location_city

            # 4) fetch weather for the report date
            today_str = date.today().isoformat()
            d_str     = report_date.isoformat()
            weather   = {}
            if d_str < today_str:
                resp    = requests.get(GS_URL, params={'city': city, 'date': d_str}, timeout=10)
                results = resp.json().get('results', [])
                if results:
                    rec = results[0]
                    weather = {
                        'temp':      rec.get('temp', 0),
                        'condition': rec.get('condition', ''),
                        'wind':      rec.get('wind', 0),
                        'humidity':  rec.get('humidity', 0),
                    }
            elif d_str == today_str:
                resp = requests.get(
                    'https://api.openweathermap.org/data/2.5/weather',
                    params={'q': city, 'appid': OWM_API_KEY, 'units': 'metric'},
                    timeout=10
                )
                j = resp.json()
                weather = {
                    'temp':      j['main']['temp'],
                    'condition': j['weather'][0]['description'],
                    'wind':      j['wind']['speed'],
                    'humidity':  j['main']['humidity'],
                }
            else:
                resp = requests.get(
                    'https://api.openweathermap.org/data/2.5/forecast',
                    params={'q': city, 'appid': OWM_API_KEY, 'units': 'metric'},
                    timeout=10
                )
                j     = resp.json()
                slot  = next((i for i in j['list'] if i['dt_txt'].startswith(d_str + ' 12:00:00')), None) or j['list'][0]
                weather = {
                    'temp':      slot['main']['temp'],
                    'condition': slot['weather'][0]['description'],
                    'wind':      slot['wind']['speed'],
                    'humidity':  slot['main']['humidity'],
                }

            # 5) helper to parse & quantize decimals
            def parse_decimal(val, field_name):
                try:
                    d  = Decimal(str(val))
                    dp = DailyReport._meta.get_field(field_name).decimal_places
                    return d.quantize(Decimal(f'1e-{dp}'))
                except Exception:
                    return Decimal('0')

            # 6) create main DailyReport with current user
            report = DailyReport.objects.create(
                project=project,
                created_by=request.user,
                date=report_date,
                temp_c=      parse_decimal(weather.get('temp', 0),      'temp_c'),
                condition=   weather.get('condition', ''),
                wind_kmh=    parse_decimal(weather.get('wind', 0),      'wind_kmh'),
                humidity_pct=parse_decimal(weather.get('humidity', 0), 'humidity_pct'),
            )

            # 7) helper to parse ints
            def parse_int(v):
                try:
                    return int(v)
                except:
                    return 0

            # 8) create related entries
            for role, qty, hrs, com in data.get('manpower', []):
                Manpower.objects.create(
                    daily_report=report,
                    role=role,
                    quantity=parse_int(qty),
                    hours=parse_int(hrs),
                    comments=com,
                )
            for eq, qty, hrs, com in data.get('equipment', []):
                Equipment.objects.create(
                    daily_report=report,
                    equipment=eq,
                    quantity=parse_int(qty),
                    hours_used=parse_int(hrs),
                    comments=com,
                )
            for act, qty, unit, desc in data.get('activities', []):
                Activity.objects.create(
                    daily_report=report,
                    activity=act,
                    quantity=parse_int(qty),
                    unit=unit,
                    description=desc,
                )
            # 8) create related entries
            for f in request.FILES.getlist('photos'):
                Photo.objects.create(daily_report=report, photo=f)

            # ── start support email notification ──
            try:
                uploader = request.user.get_full_name() or request.user.username

                # link to the full history page for this project
                reports_url = request.build_absolute_uri(
                    reverse('reportes:dailyreport_list', args=[project.pk])
                )

                # build plain-text body
                lines = [
                    f"New Daily Report Filed by {uploader}",
                    f"Project       : {project.name}",
                    f"Date          : {report.date}",
                    "",
                    "Weather:",
                    f"  • Temp       : {report.temp_c} °C",
                    f"  • Condition  : {report.condition}",
                    f"  • Wind       : {report.wind_kmh} km/h",
                    f"  • Humidity   : {report.humidity_pct}%",
                    "",
                    "Manpower Usage:",
                ]
                for m in report.manpower.all():
                    lines.append(
                        f"  - {m.role}: qty={m.quantity}, hrs={m.hours}, comments={m.comments or '—'}"
                    )
                lines += ["", "Equipment Usage:"]
                for e in report.equipment.all():
                    lines.append(
                        f"  - {e.equipment}: qty={e.quantity}, hrs={e.hours_used}, comments={e.comments or '—'}"
                    )
                lines += ["", "Activities:"]
                for a in report.activities.all():
                    lines.append(
                        f"  - {a.activity}: qty={a.quantity} {a.unit}, desc={a.description or '—'}"
                    )

                # add photos
                lines += ["", "Photos:"]
                for p in report.photos.all():
                    photo_url = request.build_absolute_uri(p.photo.url)
                    lines.append(f"  - {p.photo.name}: {photo_url}")

                # add link to full history
                lines += ["", f"View all past reports for this project: {reports_url}"]
                plain = "\n".join(lines)

                # build HTML body
                html = f"""
<html>
  <body style="font-family:Arial,sans-serif;color:#333;max-width:600px;margin:auto">
    <h2 style="color:#0d3b66;">New Daily Report by {uploader}</h2>
    <!-- Project & date -->
    <h3>Project & Report Date</h3>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1em">
      <tr>
        <th style="text-align:left;padding:6px;border:1px solid #ddd">Project</th>
        <td style="padding:6px;border:1px solid #ddd">{project.name}</td>
      </tr>
      <tr>
        <th style="text-align:left;padding:6px;border:1px solid #ddd">Date</th>
        <td style="padding:6px;border:1px solid #ddd">{report.date}</td>
      </tr>
    </table>

    <!-- Weather -->
    <h3>Weather</h3>
    <ul>
      <li><strong>Temp:</strong> {report.temp_c} °C</li>
      <li><strong>Condition:</strong> {report.condition}</li>
      <li><strong>Wind:</strong> {report.wind_kmh} km/h</li>
      <li><strong>Humidity:</strong> {report.humidity_pct}%</li>
    </ul>

    <!-- Manpower -->
    <h3>Manpower</h3>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1em">
      <thead style="background:#f0f0f0">
        <tr>
          <th style="padding:6px;border:1px solid #ddd">Role</th>
          <th style="padding:6px;border:1px solid #ddd">Qty</th>
          <th style="padding:6px;border:1px solid #ddd">Hrs</th>
          <th style="padding:6px;border:1px solid #ddd">Comments</th>
        </tr>
      </thead>
      <tbody>
"""
                for m in report.manpower.all():
                    html += (
                        "<tr>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{m.role}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{m.quantity}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{m.hours}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{m.comments or '—'}</td>"
                        "</tr>"
                    )
                html += """
      </tbody>
    </table>

    <!-- Equipment -->
    <h3>Equipment</h3>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1em">
      <thead style="background:#f0f0f0">
        <tr>
          <th style="padding:6px;border:1px solid #ddd">Item</th>
          <th style="padding:6px;border:1px solid #ddd">Qty</th>
          <th style="padding:6px;border:1px solid #ddd">Hrs Used</th>
          <th style="padding:6px;border:1px solid #ddd">Comments</th>
        </tr>
      </thead>
      <tbody>
"""
                for e in report.equipment.all():
                    html += (
                        "<tr>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{e.equipment}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{e.quantity}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{e.hours_used}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{e.comments or '—'}</td>"
                        "</tr>"
                    )
                html += """
      </tbody>
    </table>

    <!-- Activities -->
    <h3>Activities</h3>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1em">
      <thead style="background:#f0f0f0">
        <tr>
          <th style="padding:6px;border:1px solid #ddd">Activity</th>
          <th style="padding:6px;border:1px solid #ddd">Qty</th>
          <th style="padding:6px;border:1px solid #ddd">Unit</th>
          <th style="padding:6px;border:1px solid #ddd">Description</th>
        </tr>
      </thead>
      <tbody>
"""
                for a in report.activities.all():
                    html += (
                        "<tr>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{a.activity}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{a.quantity}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{a.unit}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{a.description or '—'}</td>"
                        "</tr>"
                    )
                html += f"""
      </tbody>
    </table>

    <!-- Photos -->
    <h3>Photos</h3>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1em">
      <thead style="background:#f0f0f0">
        <tr>
          <th style="padding:6px;border:1px solid #ddd">File</th>
          <th style="padding:6px;border:1px solid #ddd">View Link</th>
        </tr>
      </thead>
      <tbody>
"""
                for p in report.photos.all():
                    photo_url = request.build_absolute_uri(p.photo.url)
                    html += (
                        "<tr>"
                        f"<td style='padding:6px;border:1px solid #ddd'>{p.photo.name}</td>"
                        f"<td style='padding:6px;border:1px solid #ddd'>"
                          f"<a href='{photo_url}'>Download</a>"
                        "</td>"
                        "</tr>"
                    )
                html += f"""
      </tbody>
    </table>

    <!-- Link to full history -->
    <p style="margin-top:1em">
      <a href="{reports_url}">View full daily‐report history for this project</a>
    </p>

    <p style="margin-top:2em;font-size:.9rem;color:#666">
      Submitted by {uploader}
    </p>
  </body>
</html>
"""

                gas(
                    to="customer-support@borbollagroup.com",
                    subject=f"[New Report] {project.name} @ {report.date}",
                    body=plain,
                    html_body=html
                )
            except Exception:
                logger.exception("Failed to notify support of report %s", report.pk)
            # ── end support email notification ──

            # 9) redirect to success page
            return redirect('reportes:daily_report_success')


           

        except Exception as e:
            logger.exception("Error processing daily report: %s", e)
            return HttpResponseServerError(str(e))

    # GET → render the form
    try:
        today    = date.today()
        projects = Project.objects.filter(
            Q(start_date__lte=today),
            Q(end_date__gte=today) | Q(end_date__isnull=True)
        ).order_by('-start_date')

        return render(request, 'reportes/dailyreport_form.html', {
            'projects':     projects,
            'default_date': today.isoformat(),
        })

    except Exception as e:
        logger.exception("Error rendering daily report form: %s", e)
        return HttpResponseServerError(str(e))



class ProjectListView(ListView):
    model = Project
    template_name = 'reportes/project_list.html'
    context_object_name = 'projects'

    def get_queryset(self):
        today = timezone.localdate()
        return (
            Project.objects
                   .filter(end_date__gte=today)      # only active
                   .order_by('end_date', 'start_date')  # newest-ending first
        )


from .forms import ProjectForm, DrawingFormSet

class ProjectCreateView(CreateView):
    model = Project
    form_class = ProjectForm
    template_name = 'reportes/project_form.html'
    success_url = reverse_lazy('reportes:project_list')

    def get_context_data(self, **kw):
        ctx = super().get_context_data(**kw)
        ctx['drawings'] = DrawingFormSet(prefix='drawings')
        return ctx

    def form_valid(self, form):
        resp = super().form_valid(form)
        ds = DrawingFormSet(self.request.POST, self.request.FILES,
                            instance=self.object, prefix='drawings')
        if ds.is_valid():
            ds.save()
        # notify all contacts that a new project was created
        threading.Thread(
            target=self._notify_contacts,
            args=(self.object, True),
            daemon=True
        ).start()
        return resp
    def _notify_contacts(self, project, created: bool):
        """
        Email all Contacts of project.client in one shot,
        now including links to each DailyReport for that project.
        """
        client = project.client
        contacts_qs = client.contacts.filter(email__isnull=False).exclude(email='')
        emails = list(contacts_qs.values_list('email', flat=True))
        if not emails:
            return

        action = "created" if created else "updated"
        subject = f"Project {action.title()}: {project.name}"

        # ——— Build client + project text sections ———
        client_txt = (
            f"Client Name   : {client.legal_name}\n"
            f"Tax ID        : {client.tax_id}\n"
            f"Email         : {client.email or '—'}\n"
            f"Phone         : {client.phone or '—'}\n\n"
        )
        proj_txt = (
            f"Project Name  : {project.name}\n"
            f"Contract #    : {project.contract_number or '—'}\n"
            f"Start Date    : {project.start_date or 'TBD'}\n"
            f"End Date      : {project.end_date or 'TBD'}\n"
            f"Location City : {project.location_city}\n\n"
        )

        # ——— Build drawings text ———
        drawings = project.drawings.all()
        if drawings:
            drawings_txt = "Drawings:\n" + "\n".join(
                f" - {d.description or d.file.name}: "
                f"{self.request.build_absolute_uri(d.file.url)}"
                for d in drawings
            ) + "\n\n"
        else:
            drawings_txt = "No drawings uploaded.\n\n"

        # ——— Build daily‐reports text ———
        reports = project.dailyreport_set.order_by('-date')
        if reports.exists():
            dr_txt = "Daily Reports:\n" + "\n".join(
                f" - {r.date:%Y-%m-%d}: "
                f"{self.request.build_absolute_uri(reverse('reportes:dailyreport_detail', args=[r.pk]))}"
                for r in reports
            ) + "\n\n"
        else:
            dr_txt = "No daily reports yet.\n\n"

        # ——— Plain‐text body ———
        plain = (
            "Hello,\n\n"
            f"The project “{project.name}” has just been {action}.\n\n"
            "=== Client Information ===\n"
            + client_txt +
            "=== Project Information ===\n"
            + proj_txt +
            drawings_txt +
            dr_txt +
            "Best regards,\n"
            "Borbolla Automation Team\n"
        )

        # ——— Build styled HTML body ———
        details_html = f"""
        <html><head><meta charset="utf-8"><title>Project {action.title()}: {project.name}</title></head>
        <body style="font-family:Arial,sans-serif;color:#333;margin:0;padding:0;">
          <div style="max-width:600px;margin:auto;padding:20px;">
            <h2 style="color:#0d3b66;">Project “{project.name}” has been <em>{action}</em></h2>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Client Information</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <tr><th style="text-align:left;padding:8px;border:1px solid #ddd;">Field</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Value</th></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Name</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.legal_name}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Tax ID</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.tax_id}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Email</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.email or '—'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Phone</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.phone or '—'}</td></tr>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Project Details</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <tr><th style="text-align:left;padding:8px;border:1px solid #ddd;">Field</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Value</th></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Name</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.name}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Contract #</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.contract_number or '—'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Start Date</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.start_date or 'TBD'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">End Date</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.end_date or 'TBD'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Location City</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.location_city}</td></tr>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Drawings</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <thead style="background:#f0f0f0;">
                <tr>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Description</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Download Link</th>
                </tr>
              </thead><tbody>
        """
        if drawings:
            for d in drawings:
                url = self.request.build_absolute_uri(d.file.url)
                details_html += (
                    f"<tr>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{d.description or d.file.name}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>"
                    f"<a href='{url}'>Download</a></td>"
                    f"</tr>"
                )
        else:
            details_html += (
                "<tr>"
                "<td colspan='2' style='padding:8px;border:1px solid #ddd;"
                "text-align:center;'>No drawings uploaded.</td>"
                "</tr>"
            )

        # Daily Reports section
        details_html += """
            </tbody>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Daily Reports</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <thead style="background:#f0f0f0;">
                <tr>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Date</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Link</th>
                </tr>
              </thead><tbody>
        """
        if reports.exists():
            for r in reports:
                rpt_url = self.request.build_absolute_uri(
                    reverse('reportes:dailyreport_detail', args=[r.pk])
                )
                details_html += (
                    f"<tr>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{r.date:%Y-%m-%d}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>"
                    f"<a href='{rpt_url}'>View Report</a></td>"
                    f"</tr>"
                )
        else:
            details_html += (
                "<tr>"
                "<td colspan='2' style='padding:8px;border:1px solid #ddd;"
                "text-align:center;'>No daily reports yet.</td>"
                "</tr>"
            )
        details_html += """
              </tbody>
            </table>

            <hr style="margin-top:30px;"/>
            <p>Best regards,<br/>Borbolla Automation Team</p>
          </div>
        </body></html>
        """

        # ——— Send one email ———
        try:
            gas(
                to=emails,
                subject=subject,
                body=plain,
                html_body=details_html,
                cc="customer-support@borbollagroup.com"
            )
        except Exception:
            logger.exception(
                "Failed to send project-notification email for project %s to %s",
                project.pk, emails
            )



class ProjectUpdateView(UpdateView):
    model = Project
    form_class = ProjectForm
    template_name = 'reportes/project_form.html'
    success_url = reverse_lazy('reportes:project_list')

    def get_context_data(self, **kw):
        ctx = super().get_context_data(**kw)
        if self.request.method=='POST':
            ctx['drawings'] = DrawingFormSet(self.request.POST, self.request.FILES,
                                             instance=self.object, prefix='drawings')
        else:
            ctx['drawings'] = DrawingFormSet(instance=self.object, prefix='drawings')
        return ctx

    def form_valid(self, form):
        resp = super().form_valid(form)
        ds = DrawingFormSet(self.request.POST, self.request.FILES,
                            instance=self.object, prefix='drawings')
        if ds.is_valid():
            ds.save()
        # notify all contacts that a new project was created
        threading.Thread(
            target=self._notify_contacts,
            args=(self.object, False),
            daemon=True
        ).start()
        return resp

    def _notify_contacts(self, project, created: bool):
        """
        Email all Contacts of project.client in one shot,
        now including links to each DailyReport for that project.
        """
        client = project.client
        contacts_qs = client.contacts.filter(email__isnull=False).exclude(email='')
        emails = list(contacts_qs.values_list('email', flat=True))
        if not emails:
            return

        action = "created" if created else "updated"
        subject = f"Project {action.title()}: {project.name}"

        # ——— Build client + project text sections ———
        client_txt = (
            f"Client Name   : {client.legal_name}\n"
            f"Tax ID        : {client.tax_id}\n"
            f"Email         : {client.email or '—'}\n"
            f"Phone         : {client.phone or '—'}\n\n"
        )
        proj_txt = (
            f"Project Name  : {project.name}\n"
            f"Contract #    : {project.contract_number or '—'}\n"
            f"Start Date    : {project.start_date or 'TBD'}\n"
            f"End Date      : {project.end_date or 'TBD'}\n"
            f"Location City : {project.location_city}\n\n"
        )

        # ——— Build drawings text ———
        drawings = project.drawings.all()
        if drawings:
            drawings_txt = "Drawings:\n" + "\n".join(
                f" - {d.description or d.file.name}: "
                f"{self.request.build_absolute_uri(d.file.url)}"
                for d in drawings
            ) + "\n\n"
        else:
            drawings_txt = "No drawings uploaded.\n\n"

        # ——— Build daily‐reports text ———
        reports = project.dailyreport_set.order_by('-date')
        if reports.exists():
            dr_txt = "Daily Reports:\n" + "\n".join(
                f" - {r.date:%Y-%m-%d}: "
                f"{self.request.build_absolute_uri(reverse('reportes:dailyreport_detail', args=[r.pk]))}"
                for r in reports
            ) + "\n\n"
        else:
            dr_txt = "No daily reports yet.\n\n"

        # ——— Plain‐text body ———
        plain = (
            "Hello,\n\n"
            f"The project “{project.name}” has just been {action}.\n\n"
            "=== Client Information ===\n"
            + client_txt +
            "=== Project Information ===\n"
            + proj_txt +
            drawings_txt +
            dr_txt +
            "Best regards,\n"
            "Borbolla Automation Team\n"
        )

        # ——— Build styled HTML body ———
        details_html = f"""
        <html><head><meta charset="utf-8"><title>Project {action.title()}: {project.name}</title></head>
        <body style="font-family:Arial,sans-serif;color:#333;margin:0;padding:0;">
          <div style="max-width:600px;margin:auto;padding:20px;">
            <h2 style="color:#0d3b66;">Project “{project.name}” has been <em>{action}</em></h2>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Client Information</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <tr><th style="text-align:left;padding:8px;border:1px solid #ddd;">Field</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Value</th></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Name</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.legal_name}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Tax ID</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.tax_id}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Email</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.email or '—'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Phone</td>
                  <td style="padding:8px;border:1px solid #ddd;">{client.phone or '—'}</td></tr>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Project Details</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <tr><th style="text-align:left;padding:8px;border:1px solid #ddd;">Field</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Value</th></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Name</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.name}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Contract #</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.contract_number or '—'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Start Date</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.start_date or 'TBD'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">End Date</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.end_date or 'TBD'}</td></tr>
              <tr><td style="padding:8px;border:1px solid #ddd;">Location City</td>
                  <td style="padding:8px;border:1px solid #ddd;">{project.location_city}</td></tr>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Drawings</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <thead style="background:#f0f0f0;">
                <tr>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Description</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Download Link</th>
                </tr>
              </thead><tbody>
        """
        if drawings:
            for d in drawings:
                url = self.request.build_absolute_uri(d.file.url)
                details_html += (
                    f"<tr>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{d.description or d.file.name}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>"
                    f"<a href='{url}'>Download</a></td>"
                    f"</tr>"
                )
        else:
            details_html += (
                "<tr>"
                "<td colspan='2' style='padding:8px;border:1px solid #ddd;"
                "text-align:center;'>No drawings uploaded.</td>"
                "</tr>"
            )

        # Daily Reports section
        details_html += """
            </tbody>
            </table>

            <h3 style="border-bottom:1px solid #ddd;padding-bottom:5px;">Daily Reports</h3>
            <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
              <thead style="background:#f0f0f0;">
                <tr>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Date</th>
                  <th style="text-align:left;padding:8px;border:1px solid #ddd;">Link</th>
                </tr>
              </thead><tbody>
        """
        if reports.exists():
            for r in reports:
                rpt_url = self.request.build_absolute_uri(
                    reverse('reportes:dailyreport_detail', args=[r.pk])
                )
                details_html += (
                    f"<tr>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>{r.date:%Y-%m-%d}</td>"
                    f"<td style='padding:8px;border:1px solid #ddd;'>"
                    f"<a href='{rpt_url}'>View Report</a></td>"
                    f"</tr>"
                )
        else:
            details_html += (
                "<tr>"
                "<td colspan='2' style='padding:8px;border:1px solid #ddd;"
                "text-align:center;'>No daily reports yet.</td>"
                "</tr>"
            )
        details_html += """
              </tbody>
            </table>

            <hr style="margin-top:30px;"/>
            <p>Best regards,<br/>Borbolla Automation Team</p>
          </div>
        </body></html>
        """

        # ——— Send one email ———
        try:
            gas(
                to=emails,
                subject=subject,
                body=plain,
                html_body=details_html,
                cc="info@borbollagroup.com"
            )
        except Exception:
            logger.exception(
                "Failed to send project-notification email for project %s to %s",
                project.pk, emails
            )



class ProjectDeleteView(DeleteView):
    model = Project
    template_name = 'reportes/project_confirm_delete.html'
    success_url = reverse_lazy('reportes:project_list')


def client_list(request):
    clients = Client.objects.order_by('legal_name')
    return render(request, 'reportes/client_list.html', {
        'clients': clients,
    })

ContactFormSet = inlineformset_factory(
    Client,
    Contact,
    form=ContactForm,
    fields=['name','email','phone','role'],
    extra=1,
    can_delete=True,
)

def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        formset = ContactFormSet(request.POST, prefix='contacts')
        if form.is_valid() and formset.is_valid():
            client = form.save()
            formset.instance = client
            formset.save()
            return redirect('reportes:client_list')
    else:
        form = ClientForm()
        formset = ContactFormSet(prefix='contacts')
    return render(request, 'reportes/client_form.html', {
        'form': form,
        'formset': formset,
    })


@require_http_methods(["GET","POST"])
def client_upsert(request, pk=None):
    if request.method=="GET":
        if pk:
            client = get_object_or_404(Client, pk=pk)
        else:
            client = None
        return render(request, "reportes/client_form.html", {
            "object": client
        })

    # POST with JSON payload
    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"status":"error","error":"invalid json"}, status=400)

    # 1) upsert Client
    if pk:
        client = get_object_or_404(Client, pk=pk)
    else:
        client = Client()
    client.legal_name = payload.get("legal_name","").strip()
    client.tax_id     = payload.get("tax_id","").strip()
    client.email      = payload.get("email","").strip()
    client.phone      = payload.get("phone","").strip()
    client.save()

    # 2) handle Contacts
    sent_ids = []
    for c in payload.get("contacts",[]):
        cid = c.get("id")
        if cid:
            # update existing
            obj = Contact.objects.filter(pk=cid, client=client).first()
            if not obj:
                continue
        else:
            obj = Contact(client=client)
        # if name is empty, skip / delete
        if not c.get("name","").strip():
            if obj.pk:
                obj.delete()
            continue
        obj.name  = c["name"].strip()
        obj.email = c.get("email","").strip()
        obj.phone = c.get("phone","").strip()
        obj.role  = c.get("role","").strip()
        obj.save()
        sent_ids.append(obj.pk)

    # 3) delete any contacts the user removed
    Contact.objects.filter(client=client).exclude(pk__in=sent_ids).delete()

    return JsonResponse({"status":"ok","client_id":client.pk})

def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        form    = ClientForm(request.POST, instance=client)
        formset = ContactFormSet(request.POST, instance=client, prefix='contacts')

        if not form.is_valid() or not formset.is_valid():
            # dump errors
            logger.debug(">>> client_edit POST data: %r", request.POST)
            logger.debug(">>> form valid? %s; errors: %s", form.is_valid(), form.errors)
            logger.debug(">>> formset valid? %s; errors: %s; non_form_errors: %s",
                         formset.is_valid(), formset.errors, formset.non_form_errors())

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('reportes:client_list')

    else:
        form    = ClientForm(instance=client)
        formset = ContactFormSet(instance=client, prefix='contacts')

    return render(request, 'reportes/client_form.html', {
        'form': form, 'formset': formset, 'object': client,
    })


def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.delete()
        return redirect('reportes:client_list')
    return render(request, 'reportes/client_confirm_delete.html', {'object': client})

class FullDailyReportListView(TemplateView):
    template_name = 'reportes/full_dailyreport_list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.localdate()
        # Only projects active today, prefetch their dailyreports
        active_projects = Project.objects.filter(
            Q(start_date__lte=today),
            Q(end_date__gte=today) | Q(end_date__isnull=True)
        ).prefetch_related(
            Prefetch('dailyreport_set', to_attr='reports', queryset=DailyReport.objects.order_by('-date'))
        )
        # Prefetch those projects onto clients
        ctx['clients'] = Client.objects.order_by('legal_name').prefetch_related(
            Prefetch('projects', to_attr='active_projects', queryset=active_projects)
        )
        return ctx
    


@require_GET
def project_list_json(request):
    client_id = request.GET.get('client')
    qs = Project.objects.filter(client_id=client_id) if client_id else Project.objects.none()
    data = list(qs.values('id','name','contract_number'))
    return JsonResponse(data, safe=False)

@require_GET
def client_contacts_json(request):
    client_id = request.GET.get('client')
    qs = Contact.objects.filter(client_id=client_id) if client_id else Contact.objects.none()
    data = [{'id': c.pk, 'name': c.name, 'email': c.email} for c in qs]
    return JsonResponse(data, safe=False)
